"""
RV Trip Wizard Excel (.xlsx) parser.

Reads the 'Trip Summary' sheet and extracts:
- Trip title (row 1)
- Start date (row 2)
- Header row (row 4, anchored by 'Stop Name')
- Stop rows (row 5+, until empty rows)
"""
import hashlib
import re
from datetime import date, datetime
from typing import Optional
from dataclasses import dataclass, field, asdict

import openpyxl


@dataclass
class ParsedStop:
    """One parsed stop from the Excel file."""
    sequence: int = 0
    row_number: int = 0
    name: str = ""
    arrival_date: Optional[date] = None
    departure_date: Optional[date] = None
    nights: Optional[int] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    address: Optional[str] = None
    url: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    features_raw: Optional[str] = None
    features: list = field(default_factory=list)
    comments: Optional[str] = None
    reservation: Optional[str] = None
    miles_from_previous: Optional[float] = None
    total_miles: Optional[float] = None
    estimated_travel_time: Optional[str] = None
    camping_cost: Optional[float] = None
    meals_cost: Optional[float] = None
    misc_cost: Optional[float] = None
    fuel_cost: Optional[float] = None
    stop_total_cost: Optional[float] = None
    starting_fuel: Optional[float] = None
    fuel_used: Optional[float] = None
    arrival_fuel: Optional[float] = None
    fuel_added: Optional[float] = None
    departure_fuel: Optional[float] = None
    fingerprint: str = ""
    raw: dict = field(default_factory=dict)

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ParsedTrip:
    """Result of parsing an RV Trip Wizard Excel file."""
    title: str = ""
    start_date: Optional[str] = None
    notes: Optional[str] = None
    stops: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


# Column name normalization map
COLUMN_MAP = {
    "stop name": "name",
    "miles": "miles_from_previous",
    "total": "total_miles",
    "estimated travel time": "estimated_travel_time",
    "arrival day": "_arrival_day",
    "arrival date": "arrival_date",
    "nights": "nights",
    "departure day": "_departure_day",
    "departure date": "departure_date",
    "comments": "comments",
    "reservation number": "reservation",
    "features": "features_raw",
    "location": "address",
    "url": "url",
    "phone": "phone",
    "email": "email",
    "latitude": "latitude",
    "longitude": "longitude",
    "camping cost": "camping_cost",
    "meals cost": "meals_cost",
    "misc cost": "misc_cost",
    "fuel cost": "fuel_cost",
    "stop total cost": "stop_total_cost",
    "starting fuel": "starting_fuel",
    "fuel used": "fuel_used",
    "arrival fuel": "arrival_fuel",
    "fuel added": "fuel_added",
    "departure fuel": "departure_fuel",
}


def normalize_name(name: str) -> str:
    """Normalize a stop name for fingerprinting."""
    return re.sub(r'\s+', ' ', name.strip().lower())


def parse_date_value(val) -> Optional[date]:
    """Parse various date formats from Excel cells."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_float(val) -> Optional[float]:
    """Parse a float value, returning None for empty/invalid."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f != 0 else 0.0
    except (ValueError, TypeError):
        return None


def parse_int(val) -> Optional[int]:
    """Parse an int value."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def compute_fingerprint(stop: ParsedStop) -> str:
    """
    Deterministic fingerprint for matching stops across imports.
    Based on: normalized name + lat(5dp) + lon(5dp) + arrival + departure
    """
    parts = [
        normalize_name(stop.name),
        f"{stop.latitude:.5f}" if stop.latitude else "",
        f"{stop.longitude:.5f}" if stop.longitude else "",
        stop.arrival_date.isoformat() if stop.arrival_date else "",
        stop.departure_date.isoformat() if stop.departure_date else "",
    ]
    key = "|".join(parts)
    return hashlib.sha256(key.encode()).hexdigest()[:16]


def parse_features(raw: Optional[str]) -> list[str]:
    """Parse comma-separated features into a normalized list."""
    if not raw:
        return []
    return [f.strip() for f in raw.split(",") if f.strip()]


def parse_excel(file_path: str) -> ParsedTrip:
    """
    Parse an RV Trip Wizard Excel export file.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        ParsedTrip with title, start_date, stops, and any warnings.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = ParsedTrip()

    if "Trip Summary" not in wb.sheetnames:
        result.warnings.append("No 'Trip Summary' sheet found")
        return result

    ws = wb["Trip Summary"]

    # Row 1: Trip title
    title_cell = ws.cell(row=1, column=1).value
    if title_cell:
        result.title = str(title_cell).strip()

    # Row 2: Start date
    start_cell = ws.cell(row=2, column=1).value
    if start_cell:
        result.start_date = str(start_cell).strip()

    # Row 3: Trip notes
    notes_cell = ws.cell(row=3, column=1).value
    if notes_cell:
        notes = str(notes_cell).strip()
        if notes.startswith("Trip Notes:"):
            notes = notes[len("Trip Notes:"):].strip()
        if notes:
            result.notes = notes

    # Row 4: Headers — map column indices to field names
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val:
            key = str(val).strip().lower()
            if key in COLUMN_MAP:
                header_map[col] = COLUMN_MAP[key]

    if "name" not in header_map.values():
        result.warnings.append("Could not find 'Stop Name' header in row 4")
        return result

    # Row 5+: Stop rows
    empty_streak = 0
    sequence = 0

    for row_num in range(5, ws.max_row + 1):
        # Read all mapped values
        row_data = {}
        raw_data = {}
        for col, field_name in header_map.items():
            val = ws.cell(row=row_num, column=col).value
            raw_data[field_name] = val
            row_data[field_name] = val

        name = row_data.get("name")
        if not name or not str(name).strip():
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0
        sequence += 1

        stop = ParsedStop(
            sequence=sequence,
            row_number=row_num,
            name=str(name).strip(),
            arrival_date=parse_date_value(row_data.get("arrival_date")),
            departure_date=parse_date_value(row_data.get("departure_date")),
            nights=parse_int(row_data.get("nights")),
            latitude=parse_float(row_data.get("latitude")),
            longitude=parse_float(row_data.get("longitude")),
            address=str(row_data.get("address", "")).strip() if row_data.get("address") else None,
            url=str(row_data.get("url", "")).strip() if row_data.get("url") else None,
            phone=str(row_data.get("phone", "")).strip() if row_data.get("phone") else None,
            email=str(row_data.get("email", "")).strip() if row_data.get("email") else None,
            features_raw=str(row_data.get("features_raw", "")).strip() if row_data.get("features_raw") else None,
            comments=str(row_data.get("comments", "")).strip() if row_data.get("comments") else None,
            reservation=str(row_data.get("reservation", "")).strip() if row_data.get("reservation") else None,
            miles_from_previous=parse_float(row_data.get("miles_from_previous")),
            total_miles=parse_float(row_data.get("total_miles")),
            estimated_travel_time=str(row_data.get("estimated_travel_time", "")).strip() if row_data.get("estimated_travel_time") else None,
            camping_cost=parse_float(row_data.get("camping_cost")),
            meals_cost=parse_float(row_data.get("meals_cost")),
            misc_cost=parse_float(row_data.get("misc_cost")),
            fuel_cost=parse_float(row_data.get("fuel_cost")),
            stop_total_cost=parse_float(row_data.get("stop_total_cost")),
            starting_fuel=parse_float(row_data.get("starting_fuel")),
            fuel_used=parse_float(row_data.get("fuel_used")),
            arrival_fuel=parse_float(row_data.get("arrival_fuel")),
            fuel_added=parse_float(row_data.get("fuel_added")),
            departure_fuel=parse_float(row_data.get("departure_fuel")),
            raw={k: str(v) if v is not None else None for k, v in raw_data.items()},
        )
        stop.features = parse_features(stop.features_raw)
        stop.fingerprint = compute_fingerprint(stop)
        result.stops.append(stop)

    return result
